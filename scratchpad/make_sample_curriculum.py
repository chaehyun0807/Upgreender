from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "학사편람"

ws.merge_cells("A1:F1")
ws["A1"] = "컴퓨터공학과 교육과정 편람 (2026학년도) - 전공필수 목록"
ws["A1"].font = Font(size=14, bold=True)
ws["A1"].alignment = Alignment(horizontal="center")

header_row = 3
headers = ["과목코드", "과목명", "구분", "학점", "비고"]
for col, text in enumerate(headers, start=1):
    cell = ws.cell(row=header_row, column=col, value=text)
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="D9E1F2")
    cell.alignment = Alignment(horizontal="center")

records = [
    ("CSE101", "프로그래밍입문", "전공필수", 3, ""),
    ("CSE201-02", "자료구조", "전공필수", 3, "2023학년도 이전 입학생은 CSE201 과목으로 대체 인정"),
    ("CSE305", "운영체제", "전공필수", 3, ""),
    ("GEN101", "글쓰기와 의사소통", "필수교양", 2, ""),
    ("GEN110", "대학영어", "필수교양", 2, ""),
    ("CSE410", "인공지능", "전공선택", 3, "권장 선택과목 (필수 아님)"),  # 필수 아니므로 제외되어야 함
]
for r, row in enumerate(records, start=header_row + 1):
    for c, value in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=value)
        cell.alignment = Alignment(horizontal="center", wrap_text=False)

widths = [14, 20, 12, 8, 55]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

out_path = r"C:\Users\12cog\timetable-recommender\scratchpad\sample_curriculum.xlsx"
wb.save(out_path)
print("saved:", out_path)
