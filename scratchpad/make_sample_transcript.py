from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "성적표"

ws.merge_cells("A1:F1")
ws["A1"] = "2026학년도 성적표"
ws["A1"].font = Font(size=16, bold=True)
ws["A1"].alignment = Alignment(horizontal="center")

info_rows = [
    ("학과", "컴퓨터공학과"),
    ("학번", "20231234"),
    ("이름", "지민"),
    ("학년", "3학년"),
]
for i, (label, value) in enumerate(info_rows, start=3):
    ws[f"A{i}"] = label
    ws[f"A{i}"].font = Font(bold=True)
    ws[f"B{i}"] = value

header_row = 8
headers = ["학기", "과목코드", "과목명", "학점", "성적", "교과영역"]
for col, text in enumerate(headers, start=1):
    cell = ws.cell(row=header_row, column=col, value=text)
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="D9E1F2")
    cell.alignment = Alignment(horizontal="center")

records = [
    ("2024-1", "CSE101", "프로그래밍입문", 3, "A+", "전공필수"),
    ("2024-2", "CSE201", "자료구조", 3, "A0", "전공필수"),
    ("2025-1", "CSE314", "데이터베이스", 3, "B+", "전공선택"),
    ("2024-1", "GEN101", "글쓰기와 의사소통", 2, "A+", "필수교양"),
    ("2024-2", "COR220", "철학적 사유와 논리", 3, "B0", "중핵교양"),
    ("2025-1", "BAS150", "기초통계학", 3, "A0", "토대교양"),
    ("2024-1", "GEL140", "심리학개론", 2, "B+", "일반선택"),
]
for r, row in enumerate(records, start=header_row + 1):
    for c, value in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=value)
        cell.alignment = Alignment(horizontal="center")

total_row = header_row + len(records) + 1
ws.cell(row=total_row, column=2, value="합계").font = Font(bold=True)
ws.cell(row=total_row, column=4, value=sum(r[3] for r in records)).font = Font(bold=True)

widths = [10, 12, 22, 8, 8, 12]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

out_path = r"C:\Users\12cog\timetable-recommender\scratchpad\sample_transcript.xlsx"
wb.save(out_path)
print("saved:", out_path)
